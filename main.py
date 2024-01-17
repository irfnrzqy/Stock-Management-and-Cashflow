# List of library
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime


# dictionary for save information about items
daftar_barang = {
    "1": ["Monster", 20000, 250],
    "2": ["Taro", 37000, 250],
    "3": ["Torpedo", 28000, 250],
    "4": ["Ultra Milk", 80000, 250],
    "5": ["Beng beng", 25000, 250],
    "6": ["Nabati Coklat", 15000, 250],
    "7": ["Nabati Keju", 14000, 250],
    "8": ["Jaguar", 19000, 250],
    "9": ["Aries", 18000, 250],
    "10": ["Top", 15000, 250]
}


# hitung harga function
def hitung_harga():
    kode_barang_value = IDbarang_input.get()
    jumlah_pesanan = jumlah_entry.get()

    if kode_barang_value in daftar_barang:
        detail_barang = daftar_barang[kode_barang_value]
        harga_awal = detail_barang[1]

        if jumlah_pesanan.isdigit():
            jumlah_pesanan = int(jumlah_pesanan)

            if jumlah_pesanan <= 50:  # Kurang dari atau sama dengan 50
                total_harga = int(harga_awal) * jumlah_pesanan
            else:
                total_harga = int(harga_awal) * jumlah_pesanan
                diskon = 0.1
                total_harga -= total_harga * diskon
            return total_harga
    return 0  # Item not found, value back to 0

# print page function


def print_page():
    # delete current text in print widget
    print_text.delete(1.0, tk.END)

    # get value from kode barang
    kode_barang_value = IDbarang_input.get()

    if kode_barang_value in daftar_barang:
        detail_barang = daftar_barang[kode_barang_value]
        print_text.insert(
            tk.END, f"Kode barang: {kode_barang_value}\nNama barang: {detail_barang[0]}\nHarga: Rp. {detail_barang[1]}\nJumlah stok: {detail_barang[2]}")
    else:
        print_text.insert(tk.END, "Kode barang tidak ditemukan!")

# checkout function


def checkout():
    kode_barang_value = IDbarang_var.get()
    jumlah_pesanan = jumlah_entry.get()

    total_harga = hitung_harga()
    potongan_harga = 0  # Potongan harga awalnya diatur ke 0

    if total_harga >= 0:
        detail_barang = daftar_barang.get(kode_barang_value)
        if detail_barang:
            # Mengecek apakah stok mencukupi
            if detail_barang[2] >= int(jumlah_pesanan):
                # Mendapatkan detail transaksi
                nama_barang_checkout = detail_barang[0]
                jumlah_barang_checkout = int(jumlah_pesanan)

                # Hitung potongan harga
                if jumlah_barang_checkout > 50:
                    diskon = 0.1
                    potongan_harga = total_harga * diskon

                # Mengurangi stok barang yang dibeli
                daftar_barang[kode_barang_value][2] -= jumlah_barang_checkout

                # Total harga setelah potongan
                total_harga_checkout = total_harga - potongan_harga

                # Menulis ke spreadsheet setelah checkout
                cashflow(
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"), nama_barang_checkout, jumlah_barang_checkout, total_harga_checkout)

                # Menampilkan informasi pesanan beserta potongan harga
                output_message = f"Barang yang dibeli: {nama_barang_checkout} \nJumlah: {jumlah_barang_checkout} \nTotal harga: Rp. {total_harga_checkout} \nPotongan Harga: Rp. {potongan_harga}"
                messagebox.showinfo("Informasi Pesanan", output_message)
            else:
                messagebox.showinfo("Informasi Pesanan",
                                    "Stok barang tidak mencukupi!")
        else:
            messagebox.showinfo("Informasi Pesanan",
                                "Kode barang tidak ditemukan!")
    else:
        messagebox.showinfo(
            "Informasi Pesanan", "Kode barang tidak ditemukan atau jumlah pesanan tidak valid!")

# function for add cashflow to spreadsheet


def cashflow(tanggal, nama_barang, jumlah, total_harga):
    file_name = "cashflow.xlsx"

    try:
        # Load workbook
        workbook = load_workbook(file_name)
        sheet = workbook.active

        # Find the last row in the sheet
        last_row = sheet.max_row

        # Add new data to the next row
        sheet.cell(row=last_row + 1, column=1).value = tanggal
        sheet.cell(row=last_row + 1, column=2).value = nama_barang
        sheet.cell(row=last_row + 1, column=3).value = jumlah
        sheet.cell(row=last_row + 1, column=4).value = total_harga

        # Save changes to the file
        workbook.save(file_name)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Tanggal", "Nama Barang", "Jumlah", "Total Harga"])
        sheet.append([tanggal, nama_barang, jumlah, total_harga])

        # Save as a new file
        workbook.save(file_name)
    except Exception as e:
        print(f"Terjadi kesalahan: {str(e)}")


# GUI main script
root = tk.Tk()
root.configure(bg="white")
root.geometry("650x300")
root.resizable(False, False)
root.title("Warung Orang Aring")

input_frame = ttk.Frame(root)
input_frame.pack(padx=10, pady=5, fill="x", expand="false")

title = tk.Label(input_frame, text="Selamat datang \n di Warung Orang Aring")
title.pack(fill="both")


# Input frame
column_frame = ttk.Frame(input_frame)
column_frame.pack(fill=tk.X, padx=5, pady=5)

IDbarang_label = ttk.Label(column_frame, text="Kode Barang")
IDbarang_label.pack(side=tk.LEFT, padx=5, pady=5)

IDbarang_var = tk.StringVar()
IDbarang_input = ttk.Entry(column_frame, textvariable=IDbarang_var)
IDbarang_input.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)


# Button for print into print page
print_button = ttk.Button(input_frame, text="Cek stok", command=print_page)
print_button.pack(padx=10, pady=5, fill=tk.X, expand=True)

# text widget for print item's description
print_text = tk.Text(input_frame, height=4, width=50)
print_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# input for quantity items
column_frame2 = ttk.Frame(input_frame)
column_frame2.pack(fill=tk.X, padx=5, pady=5)

jumlah_label = ttk.Label(column_frame2, text="masukkan jumlah")
jumlah_label.pack(side=tk.LEFT, padx=5, pady=5)

jumlah_var = tk.StringVar()
jumlah_entry = ttk.Entry(column_frame2)
jumlah_entry.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

# checkout button
checkout_button = ttk.Button(
    input_frame, text="Checkout", command=checkout)
checkout_button.pack(padx=10, pady=5, fill=tk.X, expand=True)


root.mainloop()
